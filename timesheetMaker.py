import shutil
from datetime import date, timedelta
from openpyxl import load_workbook
import re

## Global Variables
def load_config(filename):
    with open(filename, "r") as config_file:
        return json.load(config_file)

config = load_config("config.json")
my_firstname = config["my_firstname"]
my_lastname = config["my_lastname"]
my_email = config["my_email"]
to_email = config["recipient_email"]
to_name = config["to_name"]


def input_day(ws, day, start, end):
    start_cell = ""
    end_cell = ""
    total_hours_cell = ""
    day = day.lower()
    if (day == "mon" or day == "monday"):
        start_cell = "C7"
        end_cell = "D7"
        total_hours_cell = "F7"
    elif (day == "tues" or day == "tuesday"):
        start_cell = "C8"
        end_cell = "D8"
        total_hours_cell = "F8"
    elif (day == "wed" or day == "wednesday"):
        start_cell = "C9"
        end_cell = "D9"
        total_hours_cell = "F9"
    elif (day == "thurs" or day == "thursday"):
        start_cell = "C10"
        end_cell = "D10"
        total_hours_cell = "F10"
    elif (day == "fri" or day == "friday"):
        start_cell = "C11"
        end_cell = "D11"
        total_hours_cell = "F11"
    elif (day == "sat" or day == "saturday"):
        start_cell = "C12"
        end_cell = "D12"
        total_hours_cell = "F12"
    
    if start_cell and end_cell:
        ws[start_cell].number_format = 'h:mm AM/PM'
        ws[end_cell].number_format = 'h:mm AM/PM'
        ws[start_cell] = start
        ws[end_cell] = end + 12/24
        
        total_hours = end + 12/24 - start
        
        ws[total_hours_cell] = format_time_as_string(total_hours)



def format_time_as_string(hours_float):
    total_minutes = round(hours_float * 24 * 60)  # Step 1: Round to the nearest minute
    hours = total_minutes // 60
    minutes = total_minutes % 60  # Step 2: Convert back to hours and minutes
    return f"{hours}h {minutes}m"



def time_to_excel_float(time_str):
    if ":" in time_str:
        hours, minutes = map(int, time_str.split(":"))
        return hours / 24.0 + (minutes / 1440.0)  # 1440 = 24 * 60
    else:
        hours = int(time_str)
        return hours / 24.0


def parse_time_str(time_str):
    match = re.search(r'(\d+)h (\d+)m', time_str)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        total_minutes = hours * 60 + minutes
        return total_minutes
    else:
        return 0  # Return 0 if the format doesn't match

def get_total_hours(ws):
    total_minutes = 0
    for cell in ["F7", "F8", "F9", "F10", "F11", "F12"]:
        cell_value = ws[cell].value
        if cell_value is not None:
            total_minutes += parse_time_str(cell_value)
    
    total_hours = total_minutes // 60
    remaining_minutes = total_minutes % 60
    return f"{total_hours}h {remaining_minutes}m"



def load_timesheet():
    original_file = config["original_file"]

    file_name = f"{(date.today() - timedelta(days=7)).strftime('%d-%m-%Y')} to {date.today().strftime('%d-%m-%Y')} Timesheet - {my_firstname} {my_lastname}.xlsx"

    shutil.copy(original_file, file_name)


    ## Now load the workbook and edit it

    wb = load_workbook(file_name) # this will open the copy
    ws = wb.active #ws = worksheet

    ## Loads in your name
    ws["B4"] = my_firstname
    ws["C4"] = my_lastname

    ws["F4"] = date.today().strftime("%d-%m-%y")

    days_worked = int(input("How many days worked? "))

    for i in range(days_worked):
        day = input("What day? ")
        start_time_str = input("Start (HH:MM or HH): ")
        end_time_str = input("End (HH:MM or HH): ")
        
        start_time_float = time_to_excel_float(start_time_str)
        end_time_float = time_to_excel_float(end_time_str)
    
        input_day(ws, day, start_time_float, end_time_float)

    ws["F14"] = get_total_hours(ws)
    wb.save(file_name)


load_timesheet()